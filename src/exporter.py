"""
Export 9-layer test cases to Excel (.xlsx) or CSV.

The Excel template mirrors the column structure used in real QA delivery,
with an added 'Layer' and 'AC Ref' column for traceability:

  | Test Case Name | Description | Steps | Expected Result |
  | Priority | Status | Layer | AC Ref | Comments |

Excel formatting:
- Header row: bold white text on dark blue fill, frozen
- Steps cell: wrapped, vertically top-aligned
- Layer column: color-coded per layer for quick visual scanning
- Priority column: red/orange/gray for High/Medium/Low
- Status column: left blank by default for QA to fill (Pass/Fail/Blocked)
- Auto-sized column widths

Public functions:
- export_xlsx(result, output_path, ticket_key=None, ticket_summary=None)
- export_csv(result, output_path)
"""

import os
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# 9 columns in this exact order. Order matters for both Excel and CSV.
COLUMNS = [
    "Test Case Name",
    "Test Case Description/Scenario",
    "Step Description",
    "Expected Result",
    "Priority",
    "Test Case Status",
    "Layer",
    "AC Ref",
    "Comments",
]

# Map from test-case dict keys to column index (0-based)
FIELD_TO_COLUMN = {
    "name": 0,
    "description": 1,
    "steps": 2,
    "expected": 3,
    "priority": 4,
    # Status (5) is intentionally left blank for QA to fill
    "layer": 6,
    "ac_ref": 7,
    "comments": 8,
}

# Color palette for the Layer column - one color per layer for quick scanning.
# Soft pastels so the cell is readable when printed.
LAYER_COLORS = {
    "UI":            "DDEBF7",  # light blue
    "Field":         "E2EFDA",  # light green
    "Conditional":   "FFF2CC",  # light yellow
    "Combination":   "FCE4D6",  # light orange
    "Action":        "EDEDED",  # light gray
    "Persistence":   "D9E1F2",  # light steel
    "Integration":   "F8CBAD",  # peach
    "Accessibility": "E4DFEC",  # light purple
    "Ambiguity":     "F4B084",  # coral - stands out
}

PRIORITY_COLORS = {
    "High":   "F8CBAD",  # peach
    "Medium": "FFF2CC",  # light yellow
    "Low":    "EDEDED",  # light gray
}

HEADER_FILL = PatternFill("solid", fgColor="305496")  # dark blue
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def export_xlsx(
    result: dict,
    output_path: str,
    ticket_key: str | None = None,
    ticket_summary: str | None = None,
) -> str:
    """
    Write the test cases to an .xlsx file matching the standard QA template.

    Args:
        result: {"test_cases": [...]}  - output of generate_test_cases()
        output_path: full path including filename, e.g. "C:/.../KAN-6_tests.xlsx"
        ticket_key: optional, written into the sheet title and a header row
        ticket_summary: optional, written into a header row above the table

    Returns:
        The output_path on success.
    """
    test_cases = result.get("test_cases", [])

    wb = Workbook()
    ws = wb.active
    sheet_title = (ticket_key or "Test Cases")[:31]  # Excel sheet titles max 31 chars
    ws.title = sheet_title

    current_row = 1

    # Metadata header rows (optional, only if ticket info provided)
    if ticket_key or ticket_summary:
        ws.cell(row=current_row, column=1, value=f"Ticket: {ticket_key or 'N/A'}")
        ws.cell(row=current_row, column=1).font = Font(name="Arial", bold=True, size=12)
        current_row += 1

        if ticket_summary:
            ws.cell(row=current_row, column=1, value=f"Summary: {ticket_summary}")
            ws.cell(row=current_row, column=1).font = Font(name="Arial", size=11)
            current_row += 1

        ws.cell(row=current_row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        ws.cell(row=current_row, column=1).font = Font(name="Arial", italic=True, size=10, color="808080")
        current_row += 1

        # Blank spacer row
        current_row += 1

    header_row = current_row

    # Write column headers
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[header_row].height = 28

    # Write test case rows
    for tc in test_cases:
        current_row += 1

        # Place values into the right columns
        for field, col_offset in FIELD_TO_COLUMN.items():
            value = tc.get(field, "")
            col_idx = col_offset + 1  # openpyxl is 1-indexed
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN_BORDER

        # Status column (col 6) - left blank, but apply border/font for visual consistency
        status_cell = ws.cell(row=current_row, column=6)
        status_cell.font = BODY_FONT
        status_cell.alignment = Alignment(horizontal="center", vertical="top")
        status_cell.border = THIN_BORDER

        # Color-code Priority cell
        priority_value = tc.get("priority", "")
        if priority_value in PRIORITY_COLORS:
            ws.cell(row=current_row, column=5).fill = PatternFill(
                "solid", fgColor=PRIORITY_COLORS[priority_value]
            )
            ws.cell(row=current_row, column=5).alignment = Alignment(
                horizontal="center", vertical="top"
            )

        # Color-code Layer cell
        layer_value = tc.get("layer", "")
        if layer_value in LAYER_COLORS:
            ws.cell(row=current_row, column=7).fill = PatternFill(
                "solid", fgColor=LAYER_COLORS[layer_value]
            )
            ws.cell(row=current_row, column=7).alignment = Alignment(
                horizontal="center", vertical="top"
            )

    # Column widths - tuned for typical content lengths
    widths = {
        1: 42,   # Test Case Name
        2: 40,   # Description
        3: 50,   # Steps
        4: 45,   # Expected
        5: 10,   # Priority
        6: 14,   # Status
        7: 14,   # Layer
        8: 30,   # AC Ref
        9: 28,   # Comments
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row so it stays visible when scrolling
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Enable auto-filter on the header row so QA can sort by Layer/Priority
    last_col_letter = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{current_row}"

    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)

    return output_path


def export_csv(result: dict, output_path: str) -> str:
    """
    Write test cases to CSV in the same column order as Excel.
    Suitable for TestRail / xRay / Zephyr import.
    """
    test_cases = result.get("test_cases", [])

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_ALL)
        writer.writerow(COLUMNS)
        for tc in test_cases:
            row = [""] * len(COLUMNS)
            for field, col_offset in FIELD_TO_COLUMN.items():
                row[col_offset] = tc.get(field, "")
            # Status column (index 5) intentionally left blank
            writer.writerow(row)

    return output_path


def default_output_path(ticket_key: str, ext: str = "xlsx") -> str:
    """
    Build a sensible default path like:
        <DEFAULT_EXPORT_DIR>/<KEY>_test_cases_<YYYYMMDD-HHMM>.<ext>
    """
    base_dir = os.environ.get("DEFAULT_EXPORT_DIR", "").strip()
    if not base_dir:
        base_dir = os.path.join(os.path.expanduser("~"), "agentic-qa-mcp", "output")

    timestamp = datetime.now().strftime("%Y%m%d-%H%M")
    safe_key = ticket_key.replace("/", "_").replace("\\", "_")
    filename = f"{safe_key}_test_cases_{timestamp}.{ext}"
    return os.path.join(base_dir, filename)


# Smoke test:  python -m src.exporter <path-to-test-cases.json> <KEY>
if __name__ == "__main__":
    import sys
    import json as jsonlib

    if len(sys.argv) < 3:
        print("Usage: python -m src.exporter <json-file> <ticket-key>", file=sys.stderr)
        sys.exit(1)

    with open(sys.argv[1], "r", encoding="utf-8") as f:
        data = jsonlib.load(f)

    key = sys.argv[2]
    xlsx_path = default_output_path(key, "xlsx")
    csv_path = default_output_path(key, "csv")

    export_xlsx(data, xlsx_path, ticket_key=key, ticket_summary="Smoke test export")
    export_csv(data, csv_path)

    print(f"[OK] Wrote {xlsx_path}", file=sys.stderr)
    print(f"[OK] Wrote {csv_path}", file=sys.stderr)